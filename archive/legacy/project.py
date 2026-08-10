"""import cv2
import face_recognition
import csv
import os
import numpy as np
from datetime import datetime

video_capture=cv2.VideoCapture(0)
jobs image face_recognition.load_image_file("photos/jobs.jpg")
jobs_encoding face recognition. face_encoding(jobs_image)[0]

ratan_tata_image face recognition.load_image_file("photos/tata.jpg")
ratan_tata_encoding face recognition. face_encoding(ratan_tata_image)[0]

sadmona_image face recognition.load_image_file("photos/sadmona.jpg")
sadmona_encoding face_recognition. face_encoding(sadmona_image)[0]

tesla_image face recognition.load_image_file("photos/tesla.jpg")
tesla_encoding face recognition. face_encoding(tesla_image)[0]

known_face_encoding =[
jobs_encoding,
ratan tata_encoding,
sadmona encoding,
tesla_encoding
]

known_faces_names - [
"Jobs",
"ratan tata",
"sadmona",
"tesla"
]

students= known faces_names.copy()
face locations =[]
face_encodings =[]
face_names =[]
s=True

now =datetime.now()
current_date= now.strftime("%Y-%m-%d") 

f open(current date:'.csv', 'w+',newline - **)
Inwriter csv.writer(f)
while True:
frame video_capture.read()
small frame cv2.resize(frame, (0,0),fx 0.25,fy 0.25)
rgb_small_frame small frame[:,1,11-1]
if s:
    face locations face recognition.face_locations(rgb_small_frame)
    face_encodings face recognition.face_encodings(rgb_small_frame, face_locations)
    face_names=[]
    for face_encoding in face_encodings:
        matches=face_recognition.compare_faces (known_face_encoding, face_encoding)
        name=""
    face distance face_recognition. face_distance (known_face_encoding, face_encoding)
    best_match_index=np.argmin(face_distance)
    if matches[best_match_index]:
        name known_faces_names[best_match_index]

    face_names.append(name)
    if name in known faces_names:
        if name in students:
            students.remove(name)
            print(students)
            current_time=now.strftime("H-XM-XS")
            Inwriter.writerow([name, current_time])

    cv2.imshow("attendence system", frame)
    if cv2.waitKey(1) & 0xFF ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
f.close()"""

"""import cv2
import face_recognition
import csv
import numpy as np
from datetime import datetime

# Start video capture
video_capture = cv2.VideoCapture(0)

# Load sample images and learn how to recognize them
jobs_image = face_recognition.load_image_file("photos/jobs.jpg")
jobs_encoding = face_recognition.face_encodings(jobs_image)[0]

ratan_tata_image = face_recognition.load_image_file("photos/tata.jpg")
ratan_tata_encoding = face_recognition.face_encodings(ratan_tata_image)[0]

sadmona_image = face_recognition.load_image_file("photos/sadmona.jpg")
sadmona_encoding = face_recognition.face_encodings(sadmona_image)[0]

tesla_image = face_recognition.load_image_file("photos/tesla.jpg")
tesla_encoding = face_recognition.face_encodings(tesla_image)[0]

# Store encodings and names
known_face_encodings = [
    jobs_encoding,
    ratan_tata_encoding,
    sadmona_encoding,
    tesla_encoding
]

known_face_names = [
    "Jobs",
    "Ratan Tata",
    "Sadmona",
    "Tesla"
]

students = known_face_names.copy()

face_locations = []
face_encodings = []
face_names = []
s = True

# Create CSV file for today's attendance
now = datetime.now()
current_date = now.strftime("%Y-%m-%d")

f = open(current_date + '.csv', 'w+', newline='')
inwriter = csv.writer(f)

while True:
    ret, frame = video_capture.read()
    if not ret:
        break

    # Resize frame for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = small_frame[:, :, ::-1]

    if s:
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        face_names = []

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = ""

            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                name = known_face_names[best_match_index]

            face_names.append(name)

            if name in known_face_names:
                if name in students:
                    students.remove(name)
                    print("Remaining students:", students)

                    current_time = now.strftime("%H:%M:%S")
                    inwriter.writerow([name, current_time])

    # Display the video feed
    cv2.imshow("Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release resources
video_capture.release()
cv2.destroyAllWindows()
f.close()"""


"""import cv2
import face_recognition
import csv
import numpy as np
from datetime import datetime

# Start video capture
video_capture = cv2.VideoCapture(0)

# Load sample images and learn how to recognize them
jobs_image = face_recognition.load_image_file("photos/jobs.jpg")

jobs_image = jobs_image.astype(np.uint8)       # force 8-bit
if jobs_image.shape[2] == 3:                   # has 3 channels
    jobs_image = jobs_image[:, :, ::-1]  

jobs_encodings = face_recognition.face_encodings(jobs_image)
if len(jobs_encodings) > 0:
    jobs_encoding = jobs_encodings[0]
else:
    raise ValueError("No face found in jobs.jpg")

ratan_tata_image = face_recognition.load_image_file("photos/tata.jpg")
ratan_tata_encoding = face_recognition.face_encodings(ratan_tata_image)[0]

sadmona_image = face_recognition.load_image_file("photos/sadmona.jpg")
sadmona_encoding = face_recognition.face_encodings(sadmona_image)[0]

tesla_image = face_recognition.load_image_file("photos/tesla.jpg")
tesla_encoding = face_recognition.face_encodings(tesla_image)[0]

# Store encodings and names
known_face_encodings = [
    jobs_encoding,
    ratan_tata_encoding,
    sadmona_encoding,
    tesla_encoding
]

known_face_names = [
    "Jobs",
    "Ratan Tata",
    "Sadmona",
    "Tesla"
]

students = known_face_names.copy()

# Create CSV file for today's attendance
current_date = datetime.now().strftime("%Y-%m-%d")
f = open(current_date + '.csv', 'w+', newline='')
inwriter = csv.writer(f)

while True:
    ret, frame = video_capture.read()
    if not ret:
        print("Failed to grab frame")
        break

    # Resize frame for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Detect faces in the frame
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []

    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = ""

        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)

        if matches[best_match_index]:
            name = known_face_names[best_match_index]

        face_names.append(name)

        if name in students:
            students.remove(name)
            print("Remaining students:", students)

            current_time = datetime.now().strftime("%H:%M:%S")
            inwriter.writerow([name, current_time])

    # Display the video feed
    cv2.imshow("Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release resources
video_capture.release()
cv2.destroyAllWindows()
f.close()"""

"""
Smart Attendance System using OpenCV + face_recognition
-------------------------------------------------------
Features:
- Opens webcam; press 'q' to quit
- Detects faces, draws rectangles
- Matches faces to known images in 'students' folder (root files or per-person subfolders)
- Writes attendance to 'attendance.xlsx' (Name, Status, Timestamp IST)
- Records each identity (including 'Unknown') only once per session to avoid duplicates

Dependencies:
    pip install opencv-python face_recognition numpy openpyxl

Notes:
- The first face found in each image is used to build that person's encoding.
- Name is taken from the filename (without extension) or the subfolder name.
"""

import cv2
import os
import sys
import time
import numpy as np
import face_recognition
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook

# ---------------- Configuration ----------------
KNOWN_FACES_DIR = "students"        # Folder with known faces
ATTENDANCE_FILE = "attendance.xlsx" # Output Excel file
TOLERANCE = 0.5                       # Lower = stricter match (0.4–0.6 is typical)
MODEL = "hog"                       # "hog" (CPU, fast) or "cnn" (GPU if dlib compiled)
FRAME_RESIZE_SCALE = 0.25           # Downscale for speed (0.25 = 1/4 size)

# India Standard Time (Asia/Kolkata) UTC+5:30
IST = timezone(timedelta(hours=5, minutes=30))

VALID_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}


def ensure_attendance_file(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Status", "Timestamp (IST)"])
        wb.save(path)


def append_attendance(path: str, name: str, status: str):
    # Append a single row and save immediately (simple & safe)
    try:
        wb = load_workbook(path)
        ws = wb.active
        ts = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
        ws.append([name, status, ts])
        wb.save(path)
    except Exception as e:
        print(f"[WARN] Could not write attendance: {e}")


def is_image_file(filename: str) -> bool:
    return os.path.splitext(filename)[1].lower() in VALID_IMAGE_EXTS


def load_known_faces(known_dir: str):
    """
    Supports two layouts:
    A) One image per person in the root:
        students/
          Ada_Lovelace.jpg
          Alan_Turing.png
    B) Multiple images per person in subfolders:
        students/
          Ada_Lovelace/1.jpg
          Ada_Lovelace/2.jpg
          Alan_Turing/1.jpg
    """
    known_encodings = []
    known_names = []

    if not os.path.isdir(known_dir):
        print(f"[ERROR] Known faces directory '{known_dir}' not found.")
        return known_encodings, known_names

    # Detect if there are subfolders
    has_subfolders = any(
        os.path.isdir(os.path.join(known_dir, entry)) for entry in os.listdir(known_dir)
    )

    if has_subfolders:
        # Pattern B: subfolders per person
        for person in sorted(os.listdir(known_dir)):
            person_dir = os.path.join(known_dir, person)
            if not os.path.isdir(person_dir):
                continue

            images = [f for f in os.listdir(person_dir) if is_image_file(f)]
            if not images:
                continue

            for img_name in images:
                img_path = os.path.join(person_dir, img_name)
                enc = encode_face_from_image(img_path)
                if enc is not None:
                    known_encodings.append(enc)
                    known_names.append(clean_name(person))
    else:
        # Pattern A: root files as one image per person
        images = [f for f in os.listdir(known_dir) if is_image_file(f)]
        for img_name in images:
            img_path = os.path.join(known_dir, img_name)
            name = clean_name(os.path.splitext(img_name)[0])
            enc = encode_face_from_image(img_path)
            if enc is not None:
                known_encodings.append(enc)
                known_names.append(name)

    print(f"[INFO] Loaded {len(known_encodings)} known face encodings.")
    return known_encodings, known_names


def clean_name(raw: str) -> str:
    # Make a nicer label from filenames/folders
    name = raw.replace("_", " ").replace("-", " ").strip()
    return " ".join(part.capitalize() for part in name.split())


def encode_face_from_image(path: str):
    try:
        image = face_recognition.load_image_file(path)
        boxes = face_recognition.face_locations(image, model=MODEL)
        if not boxes:
            print(f"[WARN] No face found in: {path}")
            return None
        encs = face_recognition.face_encodings(image, boxes)
        if not encs:
            print(f"[WARN] Could not compute encoding for: {path}")
            return None
        return encs[0]  # Use the first face
    except Exception as e:
        print(f"[WARN] Failed to process {path}: {e}")
        return None


def main():
    # Prepare attendance file
    ensure_attendance_file(ATTENDANCE_FILE)

    # Load known faces
    known_encodings, known_names = load_known_faces(KNOWN_FACES_DIR)

    # Open webcam
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)  # CAP_DSHOW helps on Windows
    if not cap.isOpened():
        print("[ERROR] Could not open webcam.")
        sys.exit(1)

    print("[INFO] Webcam started. Press 'q' to quit.")
    recorded_this_session = set()  # to avoid duplicate entries per run

    try:
        while True:
            ret, frame = cap.read()
            if not ret:
                print("[WARN] Failed to read frame from webcam.")
                break

            # Optional downscale for speed
            small_frame = cv2.resize(frame, (0, 0), fx=FRAME_RESIZE_SCALE, fy=FRAME_RESIZE_SCALE)
            rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            # Detect faces & encode
            face_locations = face_recognition.face_locations(rgb_small, model=MODEL)
            face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

            # Iterate over each face
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                name = "Unknown"

                if known_encodings:
                    distances = face_recognition.face_distance(known_encodings, face_encoding)
                    best_idx = int(np.argmin(distances))
                    best_dist = distances[best_idx]

                    if best_dist <= TOLERANCE:
                        name = known_names[best_idx]

                # Scale back up face locations since the frame we detected in was scaled
                top = int(top / FRAME_RESIZE_SCALE)
                right = int(right / FRAME_RESIZE_SCALE)
                bottom = int(bottom / FRAME_RESIZE_SCALE)
                left = int(left / FRAME_RESIZE_SCALE)

                # Draw rectangle & label
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0) if name != "Unknown" else (0, 0, 255), 2)
                label = f"{name} present" if name != "Unknown" else "Person Unknown"
                # Text background
                (text_w, text_h), baseline = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)
                cv2.rectangle(frame, (left, bottom + 5), (left + text_w + 6, bottom + 5 + text_h + baseline + 6),
                              (0, 0, 0), cv2.FILLED)
                cv2.putText(frame, label, (left + 3, bottom + text_h + 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2, cv2.LINE_AA)

                # Record attendance once per name per session
                if name not in recorded_this_session:
                    status = "present" if name != "Unknown" else "unknown"
                    append_attendance(ATTENDANCE_FILE, name, status)
                    recorded_this_session.add(name)
                    print(f"[INFO] Marked '{name}' as {status}.")

            # Show the frame
            cv2.imshow("Smart Attendance - Press 'q' to Quit", frame)

            # Exit on 'q'
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    finally:
        cap.release()
        cv2.destroyAllWindows()
        print("[INFO] Webcam closed. Goodbye!")


if __name__ == "__main__":
    main()


